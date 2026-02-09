#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
从数据库读取映射数据，生成科目映射对比Excel文件
供用户查看和修改映射结果
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
try:
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    from openpyxl.data_validation import DataValidation
from datetime import datetime
import os

def generate_mapping_excel(mapping_data, output_file, source_system_name="源系统"):
    """
    生成科目映射对比Excel文件
    
    参数:
    - mapping_data: DataFrame，包含映射数据（字段名需与中间表字段对应）
    - output_file: 输出文件路径
    - source_system_name: 源系统名称
    """
    
    # 准备Excel数据
    excel_data = {
        # 源系统科目信息（映射前）
        '源系统科目编码': mapping_data.get('source_subject_code', [''] * len(mapping_data)),
        '源系统科目名称': mapping_data.get('source_subject_name', [''] * len(mapping_data)),
        '源系统父科目编码': mapping_data.get('source_parent_code', [''] * len(mapping_data)),
        '源系统父科目名称': mapping_data.get('source_parent_name', [''] * len(mapping_data)),
        '源系统科目层级': mapping_data.get('source_subject_level', [''] * len(mapping_data)),
        '源系统科目类型': mapping_data.get('source_subject_type', [''] * len(mapping_data)),
        '源系统余额方向': mapping_data.get('source_debit_credit', [''] * len(mapping_data)),
        '源系统辅助核算': mapping_data.get('source_auxiliary_info', [''] * len(mapping_data)),
        
        # 目标系统科目信息（映射后）
        '目标系统科目编码': mapping_data.get('target_subject_code', [''] * len(mapping_data)),
        '目标系统科目名称': mapping_data.get('target_subject_name', [''] * len(mapping_data)),
        '目标系统父科目编码': mapping_data.get('target_parent_code', [''] * len(mapping_data)),
        '目标系统父科目名称': mapping_data.get('target_parent_name', [''] * len(mapping_data)),
        '目标系统科目层级': mapping_data.get('target_subject_level', [''] * len(mapping_data)),
        '目标系统科目类型': mapping_data.get('target_subject_type', [''] * len(mapping_data)),
        '目标系统余额方向': mapping_data.get('target_debit_credit', [''] * len(mapping_data)),
        '目标系统辅助核算': mapping_data.get('target_auxiliary_info', [''] * len(mapping_data)),
        
        # 匹配过程信息
        '匹配类型': mapping_data.get('match_type', [''] * len(mapping_data)),
        '匹配方法': mapping_data.get('match_method', [''] * len(mapping_data)),
        '匹配度评分': mapping_data.get('match_score', [''] * len(mapping_data)),
        '匹配置信度': mapping_data.get('match_confidence', [''] * len(mapping_data)),
        '匹配依据': mapping_data.get('match_reason', [''] * len(mapping_data)),
        
        # 处理状态
        '映射状态': mapping_data.get('mapping_status', [''] * len(mapping_data)),
        '是否已确认': mapping_data.get('is_confirmed', [''] * len(mapping_data)),
        '是否已修改': mapping_data.get('is_modified', [''] * len(mapping_data)),
        '验证结果': mapping_data.get('validation_result', [''] * len(mapping_data)),
        '是否存在冲突': mapping_data.get('conflict_flag', [''] * len(mapping_data)),
        
        # 用户操作区域（可修改）
        '用户修改目标编码': [''] * len(mapping_data),
        '用户修改目标名称': [''] * len(mapping_data),
        '用户备注': [''] * len(mapping_data),
        '用户操作': ['待处理'] * len(mapping_data),
    }
    
    df = pd.DataFrame(excel_data)
    
    # 生成Excel文件
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='映射对比表', index=False)
    
    # 使用openpyxl美化格式
    wb = load_workbook(output_file)
    ws = wb['映射对比表']
    
    # 定义样式
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    wrap_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # 设置表头样式
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = border
    
    # 设置列宽
    column_widths = {
        'A': 15, 'B': 20, 'C': 15, 'D': 20, 'E': 12, 'F': 10, 'G': 10, 'H': 15,
        'I': 15, 'J': 20, 'K': 15, 'L': 20, 'M': 12, 'N': 10, 'O': 10, 'P': 15,
        'Q': 15, 'R': 15, 'S': 12, 'T': 12, 'U': 30,
        'V': 12, 'W': 12, 'X': 12, 'Y': 12, 'Z': 12,
        'AA': 15, 'AB': 20, 'AC': 30, 'AD': 12,
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # 设置数据行样式和条件格式
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = wrap_alignment
            
            col_letter = get_column_letter(cell.column)
            if col_letter in ['E', 'F', 'G', 'M', 'N', 'O', 'S', 'T', 'V', 'W', 'X', 'Y', 'Z', 'AD']:
                cell.alignment = center_alignment
            
            # 根据状态设置背景色
            if cell.column == ws['V1'].column:  # 映射状态列
                if cell.value == '已确认' or cell.value == 'confirmed':
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif cell.value == '待确认' or cell.value == 'matched':
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            
            # 匹配置信度颜色
            if cell.column == ws['T1'].column:  # 匹配置信度列
                if cell.value == '高' or cell.value == 'high':
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif cell.value == '中' or cell.value == 'medium':
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                elif cell.value == '低' or cell.value == 'low':
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            # 用户可修改区域高亮
            if col_letter in ['AA', 'AB', 'AC', 'AD']:
                cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
                cell.font = Font(bold=True)
    
    # 添加数据验证（用户操作列）
    dv = DataValidation(type="list", formula1='"确认,拒绝,待处理,跳过"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add("AD2:AD" + str(ws.max_row))
    
    # 冻结窗格（冻结第一行和源系统信息列）
    ws.freeze_panes = 'I2'
    
    # 添加说明sheet
    add_instruction_sheet(wb, source_system_name)
    
    # 保存文件
    wb.save(output_file)
    print(f"✅ Excel文件已生成: {output_file}")
    print(f"   - 文件大小: {os.path.getsize(output_file) / 1024:.2f} KB")
    print(f"   - 包含数据: {len(df)} 条记录")
    print(f"   - 包含使用说明sheet")

def add_instruction_sheet(wb, source_system_name):
    """添加使用说明sheet"""
    ws_info = wb.create_sheet("使用说明", 0)
    info_content = [
        ["科目映射对比表 - 使用说明"],
        [""],
        ["一、表格结构说明"],
        ["1. 源系统科目信息（A-H列）：映射前的源系统科目数据，不可修改"],
        ["2. 目标系统科目信息（I-P列）：程序匹配后的目标科目，可参考"],
        ["3. 匹配过程信息（Q-U列）：匹配类型、方法、评分、依据等，不可修改"],
        ["4. 处理状态信息（V-Z列）：映射状态、确认状态等，不可修改"],
        ["5. 用户操作区域（AA-AD列）：用户可以修改的区域，灰色高亮显示"],
        [""],
        ["二、用户操作说明"],
        ["1. 查看映射结果：对比源系统和目标系统的科目信息"],
        ["2. 查看匹配依据：在'匹配依据'列查看为什么这样匹配"],
        ["3. 修改映射结果：在'用户修改目标编码'和'用户修改目标名称'列填写修改后的值"],
        ["4. 添加备注：在'用户备注'列填写说明"],
        ["5. 确认操作：在'用户操作'列选择：确认/拒绝/待处理/跳过"],
        [""],
        ["三、字段说明"],
        ["• 匹配类型：完全匹配、编码匹配、语义匹配、手动映射等"],
        ["• 匹配方法：直接匹配、大模型语义匹配、规则匹配等"],
        ["• 匹配度评分：0-100分，分数越高匹配度越高"],
        ["• 匹配置信度：高/中/低，高置信度可自动确认"],
        ["• 映射状态：待处理、已匹配、已确认、已拒绝等"],
        [""],
        ["四、操作流程"],
        ["1. 系统自动匹配后，生成此Excel文件"],
        ["2. 用户打开Excel，查看映射结果"],
        ["3. 对需要修改的记录，在用户操作区域填写修改内容"],
        ["4. 在'用户操作'列选择操作类型（确认/拒绝/待处理/跳过）"],
        ["5. 保存Excel文件，上传回系统"],
        ["6. 系统读取用户修改，更新映射关系"],
        [""],
        ["五、注意事项"],
        ["1. 用户修改目标编码/名称后，系统会验证修改的合理性"],
        ["2. 确认后的映射关系将用于后续导账处理"],
        ["3. 建议优先处理高置信度的匹配，低置信度需要仔细审核"],
        ["4. 修改后的数据需要重新验证，确保科目类型、方向等属性一致"],
        ["5. 灰色高亮区域（AA-AD列）是用户可以修改的区域"],
        [""],
        ["六、数据验证规则"],
        ["1. 目标科目编码必须在系统中存在"],
        ["2. 科目类型必须一致（资产/负债/权益/收入/费用）"],
        ["3. 余额方向建议一致（特殊情况可转换）"],
        ["4. 辅助核算需要单独处理"],
        [""],
        ["源系统：", source_system_name],
        ["生成时间：", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
    ]
    
    for i, row_data in enumerate(info_content, start=1):
        for j, value in enumerate(row_data, start=1):
            cell = ws_info.cell(row=i, column=j, value=value)
            if i == 1:
                cell.font = Font(bold=True, size=14)
            elif i in [3, 10, 17, 25, 31, 37]:
                cell.font = Font(bold=True, size=12)
    
    ws_info.column_dimensions['A'].width = 50
    ws_info.column_dimensions['B'].width = 50

if __name__ == "__main__":
    # 示例：从数据库读取数据（需要根据实际情况修改）
    # import pymysql
    # conn = pymysql.connect(host='localhost', user='user', password='pass', database='db')
    # query = "SELECT * FROM account_mapping_temp WHERE mapping_batch_id = 'BATCH001'"
    # mapping_data = pd.read_sql(query, conn)
    # conn.close()
    
    # 示例：使用示例数据
    sample_data = pd.DataFrame({
        'source_subject_code': ['1001', '1002', '1122'],
        'source_subject_name': ['库存现金', '银行存款', '应收账款'],
        'target_subject_code': ['1001', '1002', '1122'],
        'target_subject_name': ['库存现金', '银行存款', '应收账款'],
        'match_type': ['完全匹配', '完全匹配', '语义匹配（大模型）'],
        'match_method': ['直接匹配', '直接匹配', '大模型语义匹配'],
        'match_score': [100, 100, 95],
        'match_confidence': ['高', '高', '高'],
        'match_reason': ['编码和名称完全一致', '编码和名称完全一致', '大模型分析：名称和属性一致'],
        'mapping_status': ['已确认', '已确认', '待确认'],
    })
    
    output_dir = "06-输出物/其他输出"
    os.makedirs(output_dir, exist_ok=True)
    output_file = os.path.join(output_dir, "科目映射对比表_示例.xlsx")
    
    generate_mapping_excel(sample_data, output_file, "示例系统")
