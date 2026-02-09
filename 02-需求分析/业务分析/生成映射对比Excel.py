#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
生成科目映射对比Excel文件，供用户查看和修改映射结果
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

# 输出目录
output_dir = "06-输出物/其他输出"
os.makedirs(output_dir, exist_ok=True)

# 创建示例数据（实际使用时从数据库读取）
sample_data = {
    # 源系统科目信息（映射前）
    '源系统科目编码': ['1001', '1002', '1122', '150101', '160401'],
    '源系统科目名称': ['库存现金', '银行存款', '应收账款', '债券投资', '建筑工程'],
    '源系统父科目编码': ['', '', '', '1501', '1604'],
    '源系统父科目名称': ['', '', '', '长期债券投资', '在建工程'],
    '源系统科目层级': [1, 1, 1, 2, 2],
    '源系统科目类型': ['资产', '资产', '资产', '资产', '资产'],
    '源系统余额方向': ['借', '借', '借', '借', '借'],
    '源系统辅助核算': ['', '', '客户', '', ''],
    
    # 目标系统科目信息（映射后）
    '目标系统科目编码': ['1001', '1002', '1122', '150101', '160401'],
    '目标系统科目名称': ['库存现金', '银行存款', '应收账款', '债券投资', '建筑工程'],
    '目标系统父科目编码': ['', '', '', '1501', '1604'],
    '目标系统父科目名称': ['', '', '', '长期债券投资', '在建工程'],
    '目标系统科目层级': [1, 1, 1, 2, 2],
    '目标系统科目类型': ['资产', '资产', '资产', '资产', '资产'],
    '目标系统余额方向': ['借', '借', '借', '借', '借'],
    '目标系统辅助核算': ['', '', '客户', '', ''],
    
    # 匹配过程信息
    '匹配类型': ['完全匹配', '完全匹配', '语义匹配（大模型）', '语义匹配（大模型）', '语义匹配（大模型）'],
    '匹配方法': ['直接匹配', '直接匹配', '大模型语义匹配', '大模型语义匹配', '大模型语义匹配'],
    '匹配度评分': [100, 100, 95, 88, 92],
    '匹配置信度': ['高', '高', '高', '中', '高'],
    '匹配依据': ['编码和名称完全一致', '编码和名称完全一致', '大模型分析：名称和属性一致', '大模型分析：层级和语义匹配', '大模型分析：层级和语义匹配'],
    
    # 处理状态
    '映射状态': ['已确认', '已确认', '待确认', '待确认', '待确认'],
    '是否已确认': ['是', '是', '否', '否', '否'],
    '是否已修改': ['否', '否', '否', '否', '否'],
    '验证结果': ['通过', '通过', '通过', '通过', '通过'],
    '是否存在冲突': ['否', '否', '否', '否', '否'],
    
    # 用户操作区域（可修改）
    '用户修改目标编码': ['', '', '', '', ''],
    '用户修改目标名称': ['', '', '', '', ''],
    '用户备注': ['', '', '', '', ''],
    '用户操作': ['确认', '确认', '待处理', '待处理', '待处理'],
}

df = pd.DataFrame(sample_data)

# 生成Excel文件
excel_file = os.path.join(output_dir, "科目映射对比表.xlsx")
with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='映射对比表', index=False)

# 使用openpyxl美化格式
wb = load_workbook(excel_file)
ws = wb['映射对比表']

# 定义样式
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
center_alignment = Alignment(horizontal='center', vertical='center')
wrap_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

# 设置表头样式
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_alignment
    cell.border = border

# 设置列宽
column_widths = {
    'A': 15,  # 源系统科目编码
    'B': 20,  # 源系统科目名称
    'C': 15,  # 源系统父科目编码
    'D': 20,  # 源系统父科目名称
    'E': 12,  # 源系统科目层级
    'F': 10,  # 源系统科目类型
    'G': 10,  # 源系统余额方向
    'H': 15,  # 源系统辅助核算
    'I': 15,  # 目标系统科目编码
    'J': 20,  # 目标系统科目名称
    'K': 15,  # 目标系统父科目编码
    'L': 20,  # 目标系统父科目名称
    'M': 12,  # 目标系统科目层级
    'N': 10,  # 目标系统科目类型
    'O': 10,  # 目标系统余额方向
    'P': 15,  # 目标系统辅助核算
    'Q': 15,  # 匹配类型
    'R': 15,  # 匹配方法
    'S': 12,  # 匹配度评分
    'T': 12,  # 匹配置信度
    'U': 30,  # 匹配依据
    'V': 12,  # 映射状态
    'W': 12,  # 是否已确认
    'X': 12,  # 是否已修改
    'Y': 12,  # 验证结果
    'Z': 12,  # 是否存在冲突
    'AA': 15,  # 用户修改目标编码
    'AB': 20,  # 用户修改目标名称
    'AC': 30,  # 用户备注
    'AD': 12,  # 用户操作
}

for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# 设置数据行样式和条件格式
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    for cell in row:
        cell.border = border
        cell.alignment = wrap_alignment
        
        # 根据列设置不同的对齐方式
        col_letter = get_column_letter(cell.column)
        if col_letter in ['E', 'F', 'G', 'M', 'N', 'O', 'S', 'T', 'V', 'W', 'X', 'Y', 'Z', 'AD']:
            cell.alignment = center_alignment
        
        # 根据状态设置背景色
        if cell.column == ws['V1'].column:  # 映射状态列
            if cell.value == '已确认':
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif cell.value == '待确认':
                cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        
        # 匹配置信度颜色
        if cell.column == ws['T1'].column:  # 匹配置信度列
            if cell.value == '高':
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif cell.value == '中':
                cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            elif cell.value == '低':
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # 用户可修改区域高亮
        if col_letter in ['AA', 'AB', 'AC', 'AD']:
            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            cell.font = Font(bold=True)

# 冻结窗格（冻结第一行和源系统信息列）
ws.freeze_panes = 'I2'

# 添加说明sheet
ws_info = wb.create_sheet("使用说明", 0)
info_content = [
    ["科目映射对比表 - 使用说明"],
    [""],
    ["一、表格结构说明"],
    ["1. 源系统科目信息（A-H列）：映射前的源系统科目数据，不可修改"],
    ["2. 目标系统科目信息（I-P列）：程序匹配后的目标科目，可参考"],
    ["3. 匹配过程信息（Q-U列）：匹配类型、方法、评分、依据等，不可修改"],
    ["4. 处理状态信息（V-Z列）：映射状态、确认状态等，不可修改"],
    ["5. 用户操作区域（AA-AD列）：用户可以修改的区域，高亮显示"],
    [""],
    ["二、用户操作说明"],
    ["1. 查看映射结果：对比源系统和目标系统的科目信息"],
    ["2. 查看匹配依据：在'匹配依据'列查看为什么这样匹配"],
    ["3. 修改映射结果：在'用户修改目标编码'和'用户修改目标名称'列填写修改后的值"],
    ["4. 添加备注：在'用户备注'列填写说明"],
    ["5. 确认操作：在'用户操作'列选择：确认/拒绝/待处理"],
    [""],
    ["三、字段说明"],
    ["• 匹配类型：完全匹配、编码匹配、语义匹配、手动映射等"],
    ["• 匹配方法：直接匹配、大模型语义匹配、规则匹配等"],
    ["• 匹配度评分：0-100分，分数越高匹配度越高"],
    ["• 匹配置信度：高/中/低，高置信度可自动确认"],
    ["• 映射状态：待处理、已匹配、已确认、已拒绝等"],
    [""],
    ["四、注意事项"],
    ["1. 用户修改目标编码/名称后，系统会验证修改的合理性"],
    ["2. 确认后的映射关系将用于后续导账处理"],
    ["3. 建议优先处理高置信度的匹配，低置信度需要仔细审核"],
    ["4. 修改后的数据需要重新验证，确保科目类型、方向等属性一致"],
    [""],
    ["五、数据验证规则"],
    ["1. 目标科目编码必须在系统中存在"],
    ["2. 科目类型必须一致（资产/负债/权益/收入/费用）"],
    ["3. 余额方向建议一致（特殊情况可转换）"],
    ["4. 辅助核算需要单独处理"],
    [""],
    ["生成时间：", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
]

for i, row_data in enumerate(info_content, start=1):
    for j, value in enumerate(row_data, start=1):
        cell = ws_info.cell(row=i, column=j, value=value)
        if i == 1:
            cell.font = Font(bold=True, size=14)
        elif i in [3, 10, 17, 25, 31]:
            cell.font = Font(bold=True, size=12)

# 设置说明sheet列宽
ws_info.column_dimensions['A'].width = 50
ws_info.column_dimensions['B'].width = 50

# 保存文件
wb.save(excel_file)
print(f"✅ Excel文件已生成: {excel_file}")
print(f"   - 文件大小: {os.path.getsize(excel_file) / 1024:.2f} KB")
print(f"   - 包含示例数据: {len(df)} 条记录")
print(f"   - 包含使用说明sheet")
